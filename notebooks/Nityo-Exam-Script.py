import os, re, json
import fitz
from openpyxl import Workbook

def count_dominant_font(blocks):
    '''
    Iterates through all the contents of a page to identify the dominant font size, font type, and color
    This would signify the size, type, and color used for the paragraphs
    '''
    font_styles = {}
    font_sizes = {}
    font_colors = {}

    for b in blocks:  # iterate through the text blocks
        if b['type'] == 0:  # block contains text
            for l in b["lines"]:  # iterate through the text lines
                for s in l["spans"]:  # iterate through the text spans
                    # count the font_sizes usage
                    if s['size'] in font_sizes.keys():
                        font_sizes[s['size']] += 1
                    else:
                        font_sizes[s['size']] = 1
                    if s['font'] in font_styles.keys():
                        font_styles[s['font']] += 1
                    else:
                        font_styles[s['font']] = 1
                    if s['font'] in font_colors.keys():
                        font_colors[s['color']] += 1
                    else:
                        font_colors[s['color']] = 1

    max_size = max(font_sizes, key=font_sizes.get)
    max_style = max(font_styles, key=font_styles.get)
    max_color = max(font_colors, key=font_colors.get)
    #print(font_sizes)#, font_styles, font_colors)
    return max_size, max_style, max_color

def sort_blocks(block_dict):
    '''
    Sorts the json file of the blocks and generates a list containing [[X position, Y position], Blocks]
    '''
    sortedBlocks = []
    for block in block_dict['blocks']:
        
        #Test if the block contains lines of text
        #Since images block does not have lines
        try:
            block['lines']
        except:
            pass
        else:
            x = block['bbox'][0]
            y = block['bbox'][1]
            key = [x, y]
            sortedBlocks.append([key, block])
        
    sortedBlocks.sort(key = lambda k: [k[0], k[1]])
    #Sorts the list based on x (k[0]) and then y(k[1])
    
    return sortedBlocks
    #Returns the blocks without the keys

def create_paragraphs(blocks_list):
    '''
    Extracts text from the output of sort_blocks and exports it as [[X position, Y position], [Lines]]
    '''
    
    #Obtain the most used fonts
    max_size, max_style, max_color = count_dominant_font(pgdict['blocks'])
    
    #Initialize the list for storage
    paragraphs = []
    
    for block in blocks_list:
        textList = []
        for line in block[1]['lines']:
            #Test if the line contains spans
            #If none: pass, if present: enter loop
            try:
                line['spans']
            except:
                pass
            else:
                for span in line['spans']:
                #Test if the span is a paragraph content using size, color, and font
                #Since headers, page numbers, etc. have different sizes, color, and font compared to paragraph texts
                    if (span['size'] == max_size and span['color'] == max_color and span['font'] == max_style):
                        #Stores the text if it is a paragraph content
                        ##.replace, .encode, and .decode were used to remove unicode characters
                        ##.replace was needed for \xa0 since there is no space in between texts
                        textList.append(span['text'].replace('\xa0', ' ').encode("ascii", "ignore").decode())
        #Store the text list to paragraphs if there are contents
        if len(textList) > 0:
            paragraphs.append([block[0], textList])
    return paragraphs

def finalize_and_export(paragraphs):
    '''
    Joins the texts of the output from create_paragraphs and exports it in a excel file with rows and columns that matches the paragraph positions
    '''
    wb = Workbook()
    ws = wb.active
    
    #OpenPyXL uses 1 as first index
    column = 1
    row = 1
    
    #Store initial position for comparison
    x_pos = paragraphs[0][0][0]
    
    for paragraph in paragraphs:
        paragraph[1] = ''.join(paragraph[1])
        
        #Check for the block position vs paragraph
        ##All with the same x will be stored in the same column
        if (paragraph[0][0] == x_pos):
            ws.cell(row = row, column = column, value = paragraph[1])
            row += 1
        elif (paragraph[0][0] != x_pos):
            #Update row, column and x_pos if x_pos is different
            ##row = 2 so that the row will not be overwritten
            row = 2
            column += 1
            x_pos = paragraph[0][0]
            ws.cell(row = (row - 1), column = column, value = paragraph[1])
    wb.save(directory + '/../outputs/output.xlsx')

directory = os.getcwd()
file_name = '/../src_file/keppel-corporation-limited-annual-report-2018.pdf'
pdf = directory + file_name
# Initialize the directory of the file

doc = fitz.open(pdf)
pages = [11] #<--Initialize the page numbers here

for page_num in pages:

    page = doc.load_page(page_num)
    # Page 13 of pdf stored to page variable

    blocks = page.get_text('json')
    # Parse all blocks

    pgdict = json.loads(blocks)
    #Convert the blocks into json

    #Run the functions
    blocks_list = sort_blocks(pgdict)
    paragraphs = create_paragraphs(blocks_list)
    finalize_and_export(paragraphs)

